#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成三方报价单 Excel 文件
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl 库: pip install openpyxl")
    exit(1)

def create_quotation_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "三方报价单"
    
    # 设置列宽
    column_widths = {
        'A': 8,   # 序号
        'B': 20,  # 产品名称
        'C': 15,  # 产品型号
        'D': 15,  # 规格
        'E': 8,   # 单位
        'F': 10,  # 数量
        'G': 12,  # 单价
        'H': 12,  # 金额
        'I': 20,  # 备注
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    dotted_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='dotted'),
        bottom=Side(style='dotted')
    )
    
    # 标题
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = '三方报价单'
    title_cell.font = Font(name='黑体', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 表头
    headers = ['序号', '产品名称', '产品型号', '规格', '单位', '数量', '单价', '金额', '备注']
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='宋体', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 25
    
    # 数据行（3行空行）
    data_start_row = 4
    for row in range(data_start_row, data_start_row + 3):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = dotted_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
    
    # 合计行
    total_row = data_start_row + 3
    # 合并单元格：序号和产品名称列
    ws.merge_cells(f'A{total_row}:B{total_row}')
    total_label_cell = ws[f'A{total_row}']
    total_label_cell.value = '合计小写'
    total_label_cell.font = Font(name='宋体', size=11, bold=True)
    total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_label_cell.border = thin_border
    
    # 合并单元格：规格到金额列
    ws.merge_cells(f'D{total_row}:H{total_row}')
    total_amount_cell = ws[f'D{total_row}']
    total_amount_cell.value = '合计人民币金额 (大写)'
    total_amount_cell.font = Font(name='宋体', size=11)
    total_amount_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_amount_cell.border = thin_border
    
    # 备注列
    remark_cell = ws[f'I{total_row}']
    remark_cell.border = thin_border
    
    # 为合计行的所有单元格添加边框
    for col in range(1, 10):
        cell = ws.cell(row=total_row, column=col)
        if not cell.value:
            cell.border = thin_border
    ws.row_dimensions[total_row].height = 25
    
    # 备注区域
    remark_row = total_row + 1
    ws.merge_cells(f'A{remark_row}:B{remark_row}')
    remark_label_cell = ws[f'A{remark_row}']
    remark_label_cell.value = '备注:'
    remark_label_cell.font = Font(name='宋体', size=11)
    remark_label_cell.alignment = Alignment(horizontal='left', vertical='top')
    
    # 合并备注内容区域
    ws.merge_cells(f'A{remark_row}:I{remark_row}')
    ws.row_dimensions[remark_row].height = 60
    
    # 底部签名区域
    signature_row = remark_row + 2
    signature_labels = ['报价人:', '审核:', '审批:']
    signature_cols = ['A', 'D', 'G']
    
    for label, col in zip(signature_labels, signature_cols):
        cell = ws[f'{col}{signature_row}']
        cell.value = label
        cell.font = Font(name='宋体', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 为签名区域添加一些间距
    ws.row_dimensions[signature_row].height = 30
    
    # 保存文件
    filename = '三方报价单.xlsx'
    wb.save(filename)
    print(f"Excel 文件已生成: {filename}")
    return filename

if __name__ == '__main__':
    create_quotation_excel()

